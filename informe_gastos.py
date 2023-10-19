import openpyxl

# Crea una hoja Excel
wb = openpyxl.Workbook()

# Selecciona la hoja activa
sheet = wb.active

# Encabezados
sheet.cell(row=1, column=1).value = "Fecha"
sheet.cell(row=1, column=2).value = "Descripción"
sheet.cell(row=1, column=3).value = "Monto"

# Fila actual para ingresar datos
row_num = 2

# Para Ingresar la informacion que solicita
while True:
    fecha = input("Ingrese la fecha del gasto (mm-dd-aaaa): ")
    descripcion = input("Ingrese la descripción del gasto: ")
    monto = int(input("Ingrese el monto del gasto: "))

    # Escribe los datos en las filas
    sheet.cell(row=row_num, column=1).value = fecha
    sheet.cell(row=row_num, column=2).value = descripcion
    sheet.cell(row=row_num, column=3).value = monto

    # Incrementa el numero de filas para registrarlos
    row_num += 1

    # Pregunta al usuario si desea ingresar otro gasto
    respuesta = input("¿Desea ingresar otro gasto? (s/n): ")
    if respuesta != "s":
        break

# Guardar
wb.save("informe_gastos.xlsx")

#gasto más caro
max_monto = 0
max_row = 0
for row in range(2, sheet.max_row + 1):
    monto = sheet.cell(row, 3).value
    if monto > max_monto:
        max_monto = monto
        max_row = row

# gasto más bajo
min_monto = 999999999
min_row = 0
for row in range(2, sheet.max_row + 1):
    monto = sheet.cell(row, 3).value
    if monto < min_monto:
        min_monto = monto
        min_row = row

#suma gastos
total_gastos = 0
for row in range(2, sheet.max_row + 1):
    monto = sheet.cell(row, 3).value
    total_gastos += monto

print("El gasto más caro fue de $", max_monto, "el", sheet.cell(max_row, 1).value, "para", sheet.cell(max_row, 2).value)

print("El gasto más bajo fue de $", min_monto, "el", sheet.cell(min_row, 1).value, "para", sheet.cell(min_row, 2).value)

print("La suma de todos los gastos es de $", total_gastos)